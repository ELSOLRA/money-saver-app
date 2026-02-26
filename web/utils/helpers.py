from typing import Optional
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


_currency_state = {'symbol': '€', 'suffix': False}
_active_rates: dict = {}   # populated from model on startup; falls back to EXCHANGE_RATES


def set_currency_state(symbol: str, suffix: bool) -> None:
    """Update the active currency used by format_currency."""
    _currency_state['symbol'] = symbol
    _currency_state['suffix'] = suffix


def set_exchange_rates(rates: dict) -> None:
    """Store the active exchange rates used by convert_currency."""
    _active_rates.clear()
    _active_rates.update(rates)


def format_currency(amount: float) -> str:
    """Format a number as currency string using the active currency."""
    sym = _currency_state['symbol']
    if _currency_state['suffix']:
        return f"{amount:,.2f} {sym}"
    return f"{sym}{amount:,.2f}"


def format_currency_for_code(amount: float, currency_code: str) -> str:
    """Format amount in a specific currency without changing global state."""
    from utils.config import CURRENCIES, DEFAULT_CURRENCY
    cur = CURRENCIES.get(currency_code, CURRENCIES[DEFAULT_CURRENCY])
    symbol = cur['symbol']
    if cur['suffix']:
        return f"{amount:,.2f} {symbol}"
    return f"{symbol}{amount:,.2f}"


def convert_currency(amount: float, from_code: str, to_code: str) -> float:
    """Convert amount between currencies using active (or default) exchange rates."""
    if from_code == to_code:
        return amount
    if _active_rates:
        rates = _active_rates
    else:
        from utils.config import EXCHANGE_RATES
        rates = EXCHANGE_RATES
    from_rate = rates.get(from_code, 1.0)
    to_rate = rates.get(to_code, 1.0)
    return amount / from_rate * to_rate


def parse_amount(value: str) -> Optional[float]:
    """
    Parse a string value to float amount.
    Returns None if parsing fails.
    """
    try:
        # Removes common currency symbols and whitespace
        cleaned = value.strip().replace('$', '').replace('€', '').replace('kr', '').replace(',', '').replace(' ', '')
        if not cleaned:
            return None
        amount = float(cleaned)
        return amount if amount > 0 else None
    except ValueError:
        return None


def truncate_text(text: str, max_length: int = 30) -> str:
    """Truncate text with ellipsis if too long."""
    if len(text) <= max_length:
        return text
    return text[:max_length - 3] + '...'

def export_to_excel(transactions, categories, get_category_balance, get_totals, output_path: str, main_currency: str = 'EUR') -> str:
    """Export transactions to a formatted Excel file."""
    from datetime import datetime
    from utils.config import CURRENCIES
    import openpyxl.utils

    wb = openpyxl.Workbook()
    cur_sym = _currency_state['symbol']

    # Detect foreign currencies used across all transactions (sorted for stable column order)
    foreign_currencies = sorted(set(
        t.original_currency for t in transactions
        if t.original_currency and t.original_amount is not None
    ))

    # Shared styles
    header_fill = PatternFill("solid", start_color="2C3E50")
    header_font = Font(bold=True, color="FFFFFF", name="Segoe UI")
    add_fill    = PatternFill("solid", start_color="D5F5E3")
    spend_fill  = PatternFill("solid", start_color="FADBD8")
    total_fill  = PatternFill("solid", start_color="F0F0F0")

    # ── Sheet 1: All Transactions ──────────────────────────────
    ws = wb.active
    ws.title = "All Transactions"

    headers = ["#", "Date", "Time", "Category", "Action", f"Amount ({cur_sym})", "Note"]
    if foreign_currencies:
        headers += ["Orig. Currency", "Orig. Amount"]

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for i, t in enumerate(transactions, 1):
        dt = datetime.fromisoformat(t.timestamp)
        row_fill = add_fill if t.action == "add" else spend_fill
        amount = t.amount if t.action == "add" else -t.amount
        row = [i, dt.strftime("%Y-%m-%d"), dt.strftime("%H:%M:%S"),
               t.category, t.action.capitalize(), amount, t.note or ""]
        if foreign_currencies:
            row += [t.original_currency or "",
                    t.original_amount if t.original_amount is not None else ""]
        for col, val in enumerate(row, 1):
            cell = ws.cell(row=i + 1, column=col, value=val)
            cell.fill = row_fill
            cell.font = Font(name="Segoe UI", size=10)

    # ── Sheet 2: Summary (balances + totals + foreign currency columns) ─
    ws_summary = wb.create_sheet(title="Summary")

    ws_summary["A1"] = "Budget Summary"
    ws_summary["A1"].font = Font(bold=True, size=14, name="Segoe UI")

    # Build header row: fixed columns + one added/spent pair per foreign currency
    sum_headers = ["Category", f"Added ({cur_sym})", f"Spent ({cur_sym})", f"Balance ({cur_sym})"]
    for fc in foreign_currencies:
        fc_sym = CURRENCIES.get(fc, {}).get('symbol', fc)
        sum_headers += [f"Added ({fc_sym})", f"Spent ({fc_sym})"]

    for col, h in enumerate(sum_headers, 1):
        cell = ws_summary.cell(row=3, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_i, cat in enumerate(categories, start=4):
        cat_transactions = [t for t in transactions if t.category == cat]
        added   = sum(t.amount for t in cat_transactions if t.action == "add")
        spent   = sum(t.amount for t in cat_transactions if t.action == "spend")
        balance = get_category_balance(cat)
        b_fill  = add_fill if balance >= 0 else spend_fill

        ws_summary.cell(row=row_i, column=1, value=cat).font = Font(name="Segoe UI")
        ws_summary.cell(row=row_i, column=2, value=added).font = Font(name="Segoe UI", color="27AE60")
        ws_summary.cell(row=row_i, column=3, value=spent).font = Font(name="Segoe UI", color="E74C3C")
        cell = ws_summary.cell(row=row_i, column=4, value=balance)
        cell.font = Font(bold=True, name="Segoe UI")
        cell.fill = b_fill

        for fi, fc in enumerate(foreign_currencies):
            col = 5 + fi * 2
            fc_added = sum(t.original_amount for t in cat_transactions
                           if t.action == "add" and t.original_currency == fc
                           and t.original_amount is not None)
            fc_spent = sum(t.original_amount for t in cat_transactions
                           if t.action == "spend" and t.original_currency == fc
                           and t.original_amount is not None)
            if fc_added > 0:
                ws_summary.cell(row=row_i, column=col,     value=fc_added).font = Font(name="Segoe UI", color="27AE60")
            if fc_spent > 0:
                ws_summary.cell(row=row_i, column=col + 1, value=fc_spent).font = Font(name="Segoe UI", color="E74C3C")

    # Totals row
    total_row = len(categories) + 4
    ws_summary.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True, name="Segoe UI")
    ws_summary.cell(row=total_row, column=2, value=f"=SUM(B4:B{total_row-1})").font = Font(bold=True, color="27AE60")
    ws_summary.cell(row=total_row, column=3, value=f"=SUM(C4:C{total_row-1})").font = Font(bold=True, color="E74C3C")
    ws_summary.cell(row=total_row, column=4, value=f"=SUM(D4:D{total_row-1})").font = Font(bold=True)
    for col in range(1, 5):
        ws_summary.cell(row=total_row, column=col).fill = total_fill

    for fi, fc in enumerate(foreign_currencies):
        col = 5 + fi * 2
        cl_add = openpyxl.utils.get_column_letter(col)
        cl_spe = openpyxl.utils.get_column_letter(col + 1)
        cell_add = ws_summary.cell(row=total_row, column=col,
                                   value=f"=SUM({cl_add}4:{cl_add}{total_row-1})")
        cell_add.font = Font(bold=True, color="27AE60")
        cell_add.fill = total_fill
        cell_spe = ws_summary.cell(row=total_row, column=col + 1,
                                   value=f"=SUM({cl_spe}4:{cl_spe}{total_row-1})")
        cell_spe.font = Font(bold=True, color="E74C3C")
        cell_spe.fill = total_fill

    # ── Sheet 3+: Per-category transaction sheets ──────────────
    for cat in categories:
        ws_cat = wb.create_sheet(title=cat[:31])
        cat_transactions = [x for x in transactions if x.category == cat]

        # Foreign currencies used specifically in this category
        cat_foreign = sorted(set(
            t.original_currency for t in cat_transactions
            if t.original_currency and t.original_amount is not None
        ))

        cat_headers = ["Date", "Time", "Action", f"Amount ({cur_sym})", "Note"]
        if cat_foreign:
            cat_headers += ["Orig. Currency", "Orig. Amount"]

        ws_cat.append(cat_headers)
        for cell in ws_cat[1]:
            cell.fill = header_fill
            cell.font = header_font

        for t in cat_transactions:
            dt  = datetime.fromisoformat(t.timestamp)
            amt = t.amount if t.action == "add" else -t.amount
            row = [dt.strftime("%Y-%m-%d"), dt.strftime("%H:%M:%S"),
                   t.action.capitalize(), amt, t.note or ""]
            if cat_foreign:
                row += [t.original_currency or "",
                        t.original_amount if t.original_amount is not None else ""]
            ws_cat.append(row)

        last_row = ws_cat.max_row + 1
        ws_cat.cell(row=last_row, column=3, value="TOTAL").font = Font(bold=True)
        ws_cat.cell(row=last_row, column=4, value=f"=SUM(D2:D{last_row-1})").font = Font(bold=True)

    # ── Foreign currency sheets (one per currency) ──────────────
    for fc in foreign_currencies:
        fc_sym  = CURRENCIES.get(fc, {}).get('symbol', fc)
        ws_fc   = wb.create_sheet(title=f"{fc} Transactions"[:31])
        fc_headers = ["#", "Date", "Time", "Category", "Action",
                      f"Orig. Amount ({fc_sym})", f"Conv. Amount ({cur_sym})", "Note"]

        for col, h in enumerate(fc_headers, 1):
            cell = ws_fc.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        fc_transactions = [t for t in transactions
                           if t.original_currency == fc and t.original_amount is not None]

        for i, t in enumerate(fc_transactions, 1):
            dt       = datetime.fromisoformat(t.timestamp)
            row_fill = add_fill if t.action == "add" else spend_fill
            orig_amt = t.original_amount if t.action == "add" else -t.original_amount
            conv_amt = t.amount          if t.action == "add" else -t.amount
            row = [i, dt.strftime("%Y-%m-%d"), dt.strftime("%H:%M:%S"),
                   t.category, t.action.capitalize(), orig_amt, conv_amt, t.note or ""]
            for col, val in enumerate(row, 1):
                cell = ws_fc.cell(row=i + 1, column=col, value=val)
                cell.fill = row_fill
                cell.font = Font(name="Segoe UI", size=10)

        last_row = ws_fc.max_row + 1
        ws_fc.cell(row=last_row, column=5, value="TOTAL").font = Font(bold=True)
        ws_fc.cell(row=last_row, column=6, value=f"=SUM(F2:F{last_row-1})").font = Font(bold=True)
        ws_fc.cell(row=last_row, column=7, value=f"=SUM(G2:G{last_row-1})").font = Font(bold=True)

    # Auto-width for all sheets
    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=0)
            sheet.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    wb.save(output_path)
    return output_path