from typing import Optional
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


_currency_state = {'symbol': '€', 'suffix': False}
_active_rates: dict = {}   # populated from model on startup; falls back to EXCHANGE_RATES


def set_currency_state(symbol: str, suffix: bool) -> None:
    """Update the active currency used by format_currency."""
    _currency_state['symbol'] = symbol
    _currency_state['suffix'] = suffix


def set_exchange_rates(rates: dict) -> None:
    """Store the active exchange rates used by convert_currency."""
    _active_rates.clear()
    _active_rates.update(rates)


def format_currency(amount: float) -> str:
    """Format a number as currency string using the active currency."""
    sym = _currency_state['symbol']
    if _currency_state['suffix']:
        return f"{amount:,.2f} {sym}"
    return f"{sym}{amount:,.2f}"


def format_currency_for_code(amount: float, currency_code: str) -> str:
    """Format amount in a specific currency without changing global state."""
    from utils.config import CURRENCIES, DEFAULT_CURRENCY
    cur = CURRENCIES.get(currency_code, CURRENCIES[DEFAULT_CURRENCY])
    symbol = cur['symbol']
    if cur['suffix']:
        return f"{amount:,.2f} {symbol}"
    return f"{symbol}{amount:,.2f}"


def convert_currency(amount: float, from_code: str, to_code: str) -> float:
    """Convert amount between currencies using active (or default) exchange rates."""
    if from_code == to_code:
        return amount
    if _active_rates:
        rates = _active_rates
    else:
        from utils.config import EXCHANGE_RATES
        rates = EXCHANGE_RATES
    from_rate = rates.get(from_code, 1.0)
    to_rate = rates.get(to_code, 1.0)
    return amount / from_rate * to_rate


def parse_amount(value: str) -> Optional[float]:
    """
    Parse a string value to float amount.
    Returns None if parsing fails.
    """
    try:
        # Removes common currency symbols and whitespace
        cleaned = value.strip().replace('$', '').replace('€', '').replace('kr', '').replace(',', '').replace(' ', '')
        if not cleaned:
            return None
        amount = float(cleaned)
        return amount if amount > 0 else None
    except ValueError:
        return None


def truncate_text(text: str, max_length: int = 30) -> str:
    """Truncate text with ellipsis if too long."""
    if len(text) <= max_length:
        return text
    return text[:max_length - 3] + '...'

def export_to_excel(transactions, categories, get_category_balance, get_totals, output_path: str) -> str:
    """Export transactions to a formatted Excel file."""
    wb = openpyxl.Workbook()
    
    # ── Sheet 1: All Transactions ──────────────────────────────
    ws = wb.active
    ws.title = "All Transactions"

    cur_sym = _currency_state['symbol']
    headers = ["#", "Date", "Time", "Category", "Action", f"Amount ({cur_sym})", "Note"]
    header_fill = PatternFill("solid", start_color="2C3E50")
    header_font = Font(bold=True, color="FFFFFF", name="Segoe UI")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    add_fill = PatternFill("solid", start_color="D5F5E3")
    spend_fill = PatternFill("solid", start_color="FADBD8")

    for i, t in enumerate(transactions, 1):
        from datetime import datetime
        dt = datetime.fromisoformat(t.timestamp)
        row_fill = add_fill if t.action == "add" else spend_fill
        amount = t.amount if t.action == "add" else -t.amount
        row = [i, dt.strftime("%Y-%m-%d"), dt.strftime("%H:%M:%S"),
               t.category, t.action.capitalize(), amount, t.note or ""]
        for col, val in enumerate(row, 1):
            cell = ws.cell(row=i + 1, column=col, value=val)
            cell.fill = row_fill
            cell.font = Font(name="Segoe UI", size=10)

    # ── Sheet 2: Summary (balances + totals) ───────────────────
    ws_summary = wb.create_sheet(title="Summary")
    
    # Section title
    ws_summary["A1"] = "Budget Summary"
    ws_summary["A1"].font = Font(bold=True, size=14, name="Segoe UI")

    # Category balances table
    ws_summary["A3"] = "Category"
    ws_summary["B3"] = f"Added ({cur_sym})"
    ws_summary["C3"] = f"Spent ({cur_sym})"
    ws_summary["D3"] = f"Balance ({cur_sym})"
    for col in ["A3", "B3", "C3", "D3"]:
        ws_summary[col].font = Font(bold=True, color="FFFFFF", name="Segoe UI")
        ws_summary[col].fill = PatternFill("solid", start_color="2C3E50")
        ws_summary[col].alignment = Alignment(horizontal="center")

    for row_i, cat in enumerate(categories, start=4):
        cat_transactions = [t for t in transactions if t.category == cat]
        added = sum(t.amount for t in cat_transactions if t.action == "add")
        spent = sum(t.amount for t in cat_transactions if t.action == "spend")
        balance = get_category_balance(cat)
        balance_fill = PatternFill("solid", start_color="D5F5E3") if balance >= 0 else PatternFill("solid", start_color="FADBD8")

        ws_summary.cell(row=row_i, column=1, value=cat).font = Font(name="Segoe UI")
        ws_summary.cell(row=row_i, column=2, value=added).font = Font(name="Segoe UI", color="27AE60")
        ws_summary.cell(row=row_i, column=3, value=spent).font = Font(name="Segoe UI", color="E74C3C")
        cell = ws_summary.cell(row=row_i, column=4, value=balance)
        cell.font = Font(bold=True, name="Segoe UI")
        cell.fill = balance_fill

    # Totals row
    total_row = len(categories) + 4
    ws_summary.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True, name="Segoe UI")
    ws_summary.cell(row=total_row, column=2, value=f"=SUM(B4:B{total_row-1})").font = Font(bold=True, color="27AE60")
    ws_summary.cell(row=total_row, column=3, value=f"=SUM(C4:C{total_row-1})").font = Font(bold=True, color="E74C3C")
    ws_summary.cell(row=total_row, column=4, value=f"=SUM(D4:D{total_row-1})").font = Font(bold=True)
    for col in range(1, 5):
        ws_summary.cell(row=total_row, column=col).fill = PatternFill("solid", start_color="F0F0F0")

    # ── Sheet 3+: Per-category transaction sheets ──────────────
    from datetime import datetime
    for cat in categories:
        ws_cat = wb.create_sheet(title=cat[:31])
        ws_cat.append(["Date", "Time", "Action", f"Amount ({cur_sym})", "Note"])
        for cell in ws_cat[1]:
            cell.font = Font(bold=True, name="Segoe UI")
            cell.fill = PatternFill("solid", start_color="2C3E50")
            cell.font = Font(bold=True, color="FFFFFF", name="Segoe UI")

        for t in [x for x in transactions if x.category == cat]:
            dt = datetime.fromisoformat(t.timestamp)
            amt = t.amount if t.action == "add" else -t.amount
            ws_cat.append([dt.strftime("%Y-%m-%d"), dt.strftime("%H:%M:%S"),
                           t.action.capitalize(), amt, t.note or ""])

        last_row = ws_cat.max_row + 1
        ws_cat.cell(row=last_row, column=3, value="TOTAL").font = Font(bold=True)
        ws_cat.cell(row=last_row, column=4, value=f"=SUM(D2:D{last_row-1})").font = Font(bold=True)

    # Auto-width
    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=0)
            sheet.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    wb.save(output_path)
    return output_path